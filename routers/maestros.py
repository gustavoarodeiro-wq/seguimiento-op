from fastapi import APIRouter, Request, Depends, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import (
    get_db,
    ProductoTerminado, MateriaPrima, MaterialEmpaque,
    FormaFarmaceutica, EtapaProduccion,
    UnidadMedida,
)
from routers.auth import require_auth
from routers.backup import hacer_backup

router = APIRouter(prefix="/api")
templates = Jinja2Templates(directory="templates")
templates.env.cache = None  # workaround Python 3.14+

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _unidad(v):
    try:
        return UnidadMedida(v)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Unidad inválida: {v}. Usar UN, KG, L, G o ML.")


# ─────────────────────────────────────────────────────────────────────────────
# Productos Terminados
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/productos")
async def api_list_productos(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    items = db.query(ProductoTerminado).order_by(ProductoTerminado.codigo).all()
    return [_pt_dict(p) for p in items]


@router.post("/productos")
async def api_create_producto(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    body = await request.json()
    codigo = body.get("codigo", "").strip().upper()
    if not codigo:
        raise HTTPException(status_code=422, detail="El código es obligatorio.")
    if db.query(ProductoTerminado).filter(ProductoTerminado.codigo == codigo).first():
        raise HTTPException(status_code=409, detail=f"Ya existe el código '{codigo}'.")
    fid = body.get("forma_farmaceutica_id") or None
    forma_str = body.get("forma_farmaceutica", "").strip() or None
    if fid:
        ff = db.query(FormaFarmaceutica).filter(FormaFarmaceutica.id == fid).first()
        if ff:
            forma_str = ff.nombre
    cgu  = body.get("cantidad_granel_x_unidad")
    cupt = body.get("cantidad_unidades_x_pt")
    pc   = body.get("peso_comprimido")
    cxb  = body.get("cantidad_comprimidos_x_blister")
    bxpt = body.get("cantidad_blisters_x_pt")
    item = ProductoTerminado(
        codigo=codigo,
        descripcion=body.get("descripcion", "").strip(),
        unidad=_unidad(body.get("unidad", "UN")),
        forma_farmaceutica=forma_str,
        forma_farmaceutica_id=fid,
        activo=True,
        cantidad_granel_x_unidad=round(float(cgu), 1) if cgu is not None else None,
        cantidad_unidades_x_pt=int(cupt) if cupt is not None else None,
        peso_comprimido=int(pc) if pc is not None else None,
        cantidad_comprimidos_x_blister=int(cxb) if cxb is not None else None,
        cantidad_blisters_x_pt=int(bxpt) if bxpt is not None else None,
    )
    db.add(item); db.commit(); db.refresh(item)
    return _pt_dict(item)


@router.put("/productos/{item_id}")
async def api_update_producto(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    item = _get_pt(db, item_id)
    body = await request.json()
    item.descripcion = body.get("descripcion", item.descripcion).strip()
    item.unidad = _unidad(body.get("unidad", item.unidad.value))
    if "forma_farmaceutica_id" in body:
        fid = body["forma_farmaceutica_id"]
        item.forma_farmaceutica_id = fid
        if fid:
            ff = db.query(FormaFarmaceutica).filter(FormaFarmaceutica.id == fid).first()
            if ff:
                item.forma_farmaceutica = ff.nombre
        else:
            item.forma_farmaceutica = None
    else:
        item.forma_farmaceutica = body.get("forma_farmaceutica", item.forma_farmaceutica or "").strip() or None
    if "cantidad_granel_x_unidad" in body:
        cgu = body["cantidad_granel_x_unidad"]
        item.cantidad_granel_x_unidad = round(float(cgu), 1) if cgu is not None else None
    if "cantidad_unidades_x_pt" in body:
        cupt = body["cantidad_unidades_x_pt"]
        item.cantidad_unidades_x_pt = int(cupt) if cupt is not None else None
    if "peso_comprimido" in body:
        pc = body["peso_comprimido"]
        item.peso_comprimido = int(pc) if pc is not None else None
    if "cantidad_comprimidos_x_blister" in body:
        cxb = body["cantidad_comprimidos_x_blister"]
        item.cantidad_comprimidos_x_blister = int(cxb) if cxb is not None else None
    if "cantidad_blisters_x_pt" in body:
        bxpt = body["cantidad_blisters_x_pt"]
        item.cantidad_blisters_x_pt = int(bxpt) if bxpt is not None else None
    db.commit(); db.refresh(item)
    return _pt_dict(item)


@router.patch("/productos/{item_id}/toggle")
async def api_toggle_producto(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    item = _get_pt(db, item_id)
    item.activo = not item.activo
    db.commit()
    return {"id": item.id, "activo": item.activo}


def _get_pt(db, item_id):
    item = db.query(ProductoTerminado).filter(ProductoTerminado.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Producto no encontrado.")
    return item


def _pt_dict(p: ProductoTerminado):
    return {
        "id": p.id, "codigo": p.codigo, "descripcion": p.descripcion,
        "unidad": p.unidad.value, "activo": p.activo,
        "forma_farmaceutica": p.forma_farmaceutica,  # keep old string for compat
        "forma_farmaceutica_id": p.forma_farmaceutica_id,
        "forma_farmaceutica_nombre": p.forma_farmaceutica_obj.nombre if p.forma_farmaceutica_obj else p.forma_farmaceutica,
        "forma_farmaceutica_unidad": p.forma_farmaceutica_obj.unidad if p.forma_farmaceutica_obj else None,
        "granel_id": p.granel_id,
        "granel_codigo": p.granel.codigo if p.granel else None,
        "granel_descripcion": p.granel.descripcion if p.granel else None,
        "cantidad_granel": p.cantidad_granel,
        "granel_unidad": p.granel.unidad.value if p.granel else None,
        "cantidad_granel_x_unidad": p.cantidad_granel_x_unidad,
        "cantidad_unidades_x_pt": p.cantidad_unidades_x_pt,
        "peso_comprimido": p.peso_comprimido,
        "cantidad_comprimidos_x_blister": p.cantidad_comprimidos_x_blister,
        "cantidad_blisters_x_pt": p.cantidad_blisters_x_pt,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Materias Primas
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/materias-primas")
async def api_list_mp(
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    items = db.query(MateriaPrima).order_by(MateriaPrima.codigo).all()
    return [_mp_dict(m) for m in items]


CONDICIONES_VALIDAS = {"Activo", "Excipiente"}

@router.post("/materias-primas")
async def api_create_mp(
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    body = await request.json()
    codigo = body.get("codigo", "").strip().upper()
    if not codigo:
        raise HTTPException(status_code=422, detail="El código es obligatorio.")
    if db.query(MateriaPrima).filter(MateriaPrima.codigo == codigo).first():
        raise HTTPException(status_code=409, detail=f"Ya existe el código '{codigo}'.")
    condicion = body.get("condicion", "").strip() or None
    if condicion and condicion not in CONDICIONES_VALIDAS:
        raise HTTPException(status_code=422, detail="Condición inválida. Usar Activo o Excipiente.")
    unidad_mp = body.get("unidad", "KG").strip().upper()
    if unidad_mp not in UNIDADES_VALIDAS_MP:
        raise HTTPException(status_code=422, detail="Unidad inválida para materia prima. Usar G, KG o L.")
    item = MateriaPrima(
        codigo=codigo,
        descripcion=body.get("descripcion", "").strip(),
        unidad=UnidadMedida(unidad_mp),
        condicion=condicion,
        activo=True,
    )
    db.add(item); db.commit(); db.refresh(item)
    return _mp_dict(item)


@router.put("/materias-primas/{item_id}")
async def api_update_mp(
    item_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    item = _get_mp(db, item_id)
    body = await request.json()
    item.descripcion = body.get("descripcion", item.descripcion).strip()
    unidad_mp = body.get("unidad", item.unidad.value).strip().upper()
    if unidad_mp not in UNIDADES_VALIDAS_MP:
        raise HTTPException(status_code=422, detail="Unidad inválida para materia prima. Usar G, KG o L.")
    item.unidad = UnidadMedida(unidad_mp)
    condicion = body.get("condicion", item.condicion or "").strip() or None
    if condicion and condicion not in CONDICIONES_VALIDAS:
        raise HTTPException(status_code=422, detail="Condición inválida. Usar Activo o Excipiente.")
    item.condicion = condicion
    db.commit(); db.refresh(item)
    return _mp_dict(item)


@router.patch("/materias-primas/{item_id}/toggle")
async def api_toggle_mp(
    item_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    item = _get_mp(db, item_id)
    item.activo = not item.activo
    db.commit()
    return {"id": item.id, "activo": item.activo}


def _get_mp(db, item_id):
    item = db.query(MateriaPrima).filter(MateriaPrima.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Materia prima no encontrada.")
    return item


def _mp_dict(m: MateriaPrima):
    return {"id": m.id, "codigo": m.codigo, "descripcion": m.descripcion,
            "unidad": m.unidad.value, "condicion": m.condicion, "activo": m.activo}


# ─────────────────────────────────────────────────────────────────────────────
# Materiales de Empaque
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/materiales-empaque")
async def api_list_me(
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    items = db.query(MaterialEmpaque).order_by(MaterialEmpaque.codigo).all()
    return [_me_dict(m) for m in items]


@router.post("/materiales-empaque")
async def api_create_me(
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    body = await request.json()
    codigo = body.get("codigo", "").strip().upper()
    if not codigo:
        raise HTTPException(status_code=422, detail="El código es obligatorio.")
    if db.query(MaterialEmpaque).filter(MaterialEmpaque.codigo == codigo).first():
        raise HTTPException(status_code=409, detail=f"Ya existe el código '{codigo}'.")
    unidad_me = body.get("unidad", "UN").strip().upper()
    if unidad_me not in UNIDADES_VALIDAS_ME:
        raise HTTPException(status_code=422, detail="Unidad inválida. Usar UN o KG.")
    item = MaterialEmpaque(
        codigo=codigo,
        descripcion=body.get("descripcion", "").strip(),
        unidad=UnidadMedida(unidad_me),
        clasificacion=body.get("clasificacion", "").strip() or None,
        activo=True,
    )
    db.add(item); db.commit(); db.refresh(item)
    return _me_dict(item)


@router.put("/materiales-empaque/{item_id}")
async def api_update_me(
    item_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    item = _get_me(db, item_id)
    body = await request.json()
    item.descripcion = body.get("descripcion", item.descripcion).strip()
    unidad_me = body.get("unidad", item.unidad.value).strip().upper()
    if unidad_me not in UNIDADES_VALIDAS_ME:
        raise HTTPException(status_code=422, detail="Unidad inválida. Usar UN o KG.")
    item.unidad = UnidadMedida(unidad_me)
    item.clasificacion = body.get("clasificacion", item.clasificacion or "").strip() or None
    db.commit(); db.refresh(item)
    return _me_dict(item)


@router.patch("/materiales-empaque/{item_id}/toggle")
async def api_toggle_me(
    item_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    item = _get_me(db, item_id)
    item.activo = not item.activo
    db.commit()
    return {"id": item.id, "activo": item.activo}


def _get_me(db, item_id):
    item = db.query(MaterialEmpaque).filter(MaterialEmpaque.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Material de empaque no encontrado.")
    return item


def _me_dict(m: MaterialEmpaque):
    return {"id": m.id, "codigo": m.codigo, "descripcion": m.descripcion,
            "unidad": m.unidad.value, "clasificacion": m.clasificacion, "activo": m.activo}


# ─────────────────────────────────────────────────────────────────────────────
# Formas Farmacéuticas + Etapas
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/formas-farmaceuticas")
async def api_list_ff(
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    items = db.query(FormaFarmaceutica).order_by(FormaFarmaceutica.nombre).all()
    return [_ff_dict(f, db) for f in items]


@router.post("/formas-farmaceuticas")
async def api_create_ff(
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    body = await request.json()
    nombre = body.get("nombre", "").strip()
    if not nombre:
        raise HTTPException(status_code=422, detail="El nombre es obligatorio.")
    if db.query(FormaFarmaceutica).filter(FormaFarmaceutica.nombre == nombre).first():
        raise HTTPException(status_code=409, detail=f"Ya existe '{nombre}'.")
    unidad_ff = body.get("unidad", "").strip().upper() or None
    if unidad_ff and unidad_ff not in ("G", "ML"):
        raise HTTPException(status_code=422, detail="Unidad inválida para forma farmacéutica. Usar G o ML.")
    ff = FormaFarmaceutica(nombre=nombre, unidad=unidad_ff, activo=True)
    db.add(ff); db.commit(); db.refresh(ff)
    return _ff_dict(ff, db)


@router.put("/formas-farmaceuticas/{ff_id}")
async def api_update_ff(
    ff_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    ff = _get_ff(db, ff_id)
    body = await request.json()
    ff.nombre = body.get("nombre", ff.nombre).strip()
    if "unidad" in body:
        unidad_ff = body["unidad"].strip().upper() if body["unidad"] else None
        if unidad_ff and unidad_ff not in ("G", "ML"):
            raise HTTPException(status_code=422, detail="Unidad inválida para forma farmacéutica. Usar G o ML.")
        ff.unidad = unidad_ff
    db.commit(); db.refresh(ff)
    return _ff_dict(ff, db)


@router.patch("/formas-farmaceuticas/{ff_id}/toggle")
async def api_toggle_ff(
    ff_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    ff = _get_ff(db, ff_id)
    ff.activo = not ff.activo
    db.commit()
    return {"id": ff.id, "activo": ff.activo}


@router.delete("/formas-farmaceuticas/{ff_id}")
async def api_delete_ff(
    ff_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    ff = _get_ff(db, ff_id)
    total = db.query(ProductoTerminado).filter(ProductoTerminado.forma_farmaceutica_id == ff_id).count()
    if total > 0:
        raise HTTPException(status_code=409, detail=f"No se puede eliminar: {total} producto(s) usan esta forma farmacéutica.")
    db.query(EtapaProduccion).filter(EtapaProduccion.forma_farmaceutica_id == ff_id).delete()
    db.delete(ff)
    db.commit()
    return {"ok": True}


# ── Productos por forma ───────────────────────────────────────────────────────

@router.get("/formas-farmaceuticas/{ff_id}/productos")
async def api_productos_por_forma(
    ff_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    _get_ff(db, ff_id)
    productos = (
        db.query(ProductoTerminado)
        .filter(ProductoTerminado.forma_farmaceutica_id == ff_id)
        .order_by(ProductoTerminado.codigo)
        .all()
    )
    return [
        {"id": p.id, "codigo": p.codigo, "descripcion": p.descripcion, "activo": p.activo}
        for p in productos
    ]


# ── Etapas ────────────────────────────────────────────────────────────────────

@router.get("/formas-farmaceuticas/{ff_id}/etapas")
async def api_list_etapas(
    ff_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    _get_ff(db, ff_id)
    etapas = (db.query(EtapaProduccion)
              .filter(EtapaProduccion.forma_farmaceutica_id == ff_id)
              .order_by(EtapaProduccion.orden)
              .all())
    return [_etapa_dict(e) for e in etapas]


@router.post("/formas-farmaceuticas/{ff_id}/etapas")
async def api_create_etapa(
    ff_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    _get_ff(db, ff_id)
    body = await request.json()
    nombre = body.get("nombre", "").strip()
    if not nombre:
        raise HTTPException(status_code=422, detail="El nombre de la etapa es obligatorio.")
    # Siguiente número de orden
    max_orden = (db.query(EtapaProduccion)
                 .filter(EtapaProduccion.forma_farmaceutica_id == ff_id)
                 .count())
    etapa = EtapaProduccion(
        forma_farmaceutica_id=ff_id,
        orden=body.get("orden", max_orden + 1),
        nombre=nombre,
        activo=True,
    )
    db.add(etapa); db.commit(); db.refresh(etapa)
    return _etapa_dict(etapa)


@router.put("/etapas/{etapa_id}")
async def api_update_etapa(
    etapa_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    etapa = _get_etapa(db, etapa_id)
    body = await request.json()
    etapa.nombre = body.get("nombre", etapa.nombre).strip()
    etapa.orden  = body.get("orden", etapa.orden)
    db.commit(); db.refresh(etapa)
    return _etapa_dict(etapa)


@router.patch("/etapas/{etapa_id}/toggle")
async def api_toggle_etapa(
    etapa_id: int,
    request: Request, db: Session = Depends(get_db), user: dict = Depends(require_auth)
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    etapa = _get_etapa(db, etapa_id)
    etapa.activo = not etapa.activo
    db.commit()
    return {"id": etapa.id, "activo": etapa.activo}


def _get_ff(db, ff_id):
    ff = db.query(FormaFarmaceutica).filter(FormaFarmaceutica.id == ff_id).first()
    if not ff:
        raise HTTPException(status_code=404, detail="Forma farmacéutica no encontrada.")
    return ff


def _get_etapa(db, etapa_id):
    e = db.query(EtapaProduccion).filter(EtapaProduccion.id == etapa_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Etapa no encontrada.")
    return e


def _ff_dict(f: FormaFarmaceutica, db: Session):
    n_etapas = db.query(EtapaProduccion).filter(
        EtapaProduccion.forma_farmaceutica_id == f.id,
        EtapaProduccion.activo == True,
    ).count()
    total_productos = db.query(ProductoTerminado).filter(ProductoTerminado.forma_farmaceutica_id == f.id).count()
    return {"id": f.id, "nombre": f.nombre, "unidad": f.unidad, "activo": f.activo,
            "n_etapas": n_etapas, "total_productos": total_productos}


def _etapa_dict(e: EtapaProduccion):
    return {"id": e.id, "forma_farmaceutica_id": e.forma_farmaceutica_id,
            "orden": e.orden, "nombre": e.nombre, "activo": e.activo,
            "n_areas": len(e.areas)}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers Excel
# ─────────────────────────────────────────────────────────────────────────────

UNIDADES_VALIDAS    = {"UN", "KG", "L"}       # productos
UNIDADES_VALIDAS_ME = {"UN", "KG"}            # materiales de empaque
UNIDADES_VALIDAS_MP = {"G", "KG", "L"}        # materias primas

def _excel_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _estilo_encabezado(ws, cols: list[str]):
    """Aplica estilo al encabezado y ajusta el ancho de columnas."""
    fill = PatternFill("solid", fgColor="1D4ED8")
    font = Font(bold=True, color="FFFFFF")
    for i, nombre in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=nombre)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(nombre) + 4, 16)


def _leer_excel(content: bytes) -> tuple[list[str], list[list]]:
    """Lee el archivo xlsx y devuelve (encabezados_lower, filas)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo está vacío.")
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def _cel(row, headers, col) -> str:
    """Extrae el valor de una celda por nombre de columna (normalizado)."""
    try:
        idx = headers.index(col)
        v = row[idx]
        return str(v).strip() if v is not None else ""
    except (ValueError, IndexError):
        return ""

def _cel_num(row, headers, col, default=None):
    """Extrae un valor numérico de una celda; devuelve default si está vacío o ausente."""
    try:
        idx = headers.index(col)
        v = row[idx]
        if v is None or str(v).strip() == "":
            return default
        return float(str(v).strip())
    except (ValueError, IndexError):
        return default


# ─────────────────────────────────────────────────────────────────────────────
# Plantillas Excel descargables
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/productos/plantilla")
async def plantilla_productos(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    items = db.query(ProductoTerminado).order_by(ProductoTerminado.codigo).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    cols = ["codigo", "descripcion", "unidad", "forma_farmaceutica"]
    _estilo_encabezado(ws, cols)
    for p in items:
        ws.append([
            p.codigo,
            p.descripcion,
            p.unidad.value,
            p.forma_farmaceutica or "",
        ])
    if not items:
        ws.append(["PTX-001", "Ejemplo Producto", "UN", "Comprimidos"])
    return _excel_response(wb, "export_productos.xlsx")


@router.get("/productos/plantilla-comprimidos")
async def plantilla_productos_comprimidos(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    items = (
        db.query(ProductoTerminado)
        .join(FormaFarmaceutica, ProductoTerminado.forma_farmaceutica_id == FormaFarmaceutica.id, isouter=True)
        .filter(func.lower(FormaFarmaceutica.nombre).contains("comprimido"))
        .order_by(ProductoTerminado.codigo)
        .all()
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comprimidos"
    cols = ["codigo", "descripcion", "unidad", "forma_farmaceutica",
            "peso_comprimido_mg", "comprimidos_x_blister", "blisters_x_pt"]
    _estilo_encabezado(ws, cols)
    for p in items:
        ws.append([
            p.codigo,
            p.descripcion,
            p.unidad.value,
            p.forma_farmaceutica or "",
            p.peso_comprimido,
            p.cantidad_comprimidos_x_blister,
            p.cantidad_blisters_x_pt,
        ])
    if not items:
        ws.append(["PTX-001", "Ejemplo Comprimido", "UN", "Comprimidos", 500, 10, 3])
    nota = ws.cell(row=len(items) + 3, column=1,
        value="* peso_comprimido_mg: entero (mg)  |  comprimidos_x_blister y blisters_x_pt: enteros  |  "
              "forma_farmaceutica: nombre exacto del sistema")
    nota.font = Font(italic=True, color="6B7280")
    ws.merge_cells(f"A{len(items)+3}:G{len(items)+3}")
    return _excel_response(wb, "export_productos_comprimidos.xlsx")


@router.get("/productos/plantilla-liquidos")
async def plantilla_productos_liquidos(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    items = (
        db.query(ProductoTerminado)
        .join(FormaFarmaceutica, ProductoTerminado.forma_farmaceutica_id == FormaFarmaceutica.id, isouter=True)
        .filter(
            FormaFarmaceutica.unidad.isnot(None),
            ~func.lower(FormaFarmaceutica.nombre).contains("comprimido"),
        )
        .order_by(ProductoTerminado.codigo)
        .all()
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidos y Solidos"
    cols = ["codigo", "descripcion", "unidad", "forma_farmaceutica",
            "granel_x_unidad", "unidades_x_pt"]
    _estilo_encabezado(ws, cols)
    for p in items:
        ws.append([
            p.codigo,
            p.descripcion,
            p.unidad.value,
            p.forma_farmaceutica or "",
            p.cantidad_granel_x_unidad,
            p.cantidad_unidades_x_pt,
        ])
    if not items:
        ws.append(["PTX-002", "Ejemplo Suspensión", "UN", "Suspensión oral", 5.0, 1])
    nota = ws.cell(row=len(items) + 3, column=1,
        value="* granel_x_unidad: número con 1 decimal (mL o g según la FF)  |  unidades_x_pt: entero  |  "
              "forma_farmaceutica: nombre exacto del sistema")
    nota.font = Font(italic=True, color="6B7280")
    ws.merge_cells(f"A{len(items)+3}:F{len(items)+3}")
    return _excel_response(wb, "export_productos_liquidos.xlsx")


@router.get("/materias-primas/plantilla")
async def plantilla_mp(user: dict = Depends(require_auth)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materias Primas"
    cols = ["codigo", "descripcion", "unidad", "condicion"]
    _estilo_encabezado(ws, cols)
    ws.append(["MP-001", "Amoxicilina Trihidrato", "KG", "Activo"])
    nota = ws.cell(row=3, column=1, value="* unidad: G, KG o L  |  condicion: Activo o Excipiente (opcional)")
    nota.font = Font(italic=True, color="6B7280")
    ws.merge_cells("A3:D3")
    return _excel_response(wb, "plantilla_materias_primas.xlsx")


@router.get("/materiales-empaque/plantilla")
async def plantilla_me(user: dict = Depends(require_auth)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiales Empaque"
    cols = ["codigo", "descripcion", "unidad", "clasificacion"]
    _estilo_encabezado(ws, cols)
    ws.append(["ME-001", "Blíster PVC-Aluminio 10x10", "UN", "Estuche"])
    nota = ws.cell(row=3, column=1, value="* unidad: UN o KG  |  clasificacion: libre (ej: Estuche, Prospecto, Etiqueta, Frasco…)")
    nota.font = Font(italic=True, color="6B7280")
    ws.merge_cells("A3:D3")
    return _excel_response(wb, "plantilla_materiales_empaque.xlsx")


@router.get("/formas-farmaceuticas/plantilla")
async def plantilla_ff(user: dict = Depends(require_auth)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formas Farmacéuticas"
    cols = ["nombre", "etapas"]
    _estilo_encabezado(ws, cols)
    ws.append(["Comprimidos", "Pesada,Granulación,Compresión,Recubrimiento,Envasado"])
    # Nota aclaratoria
    nota = ws.cell(row=3, column=1, value="* La columna 'etapas' es opcional. Separar con comas en orden de ejecución.")
    nota.font = Font(italic=True, color="6B7280")
    ws.merge_cells("A3:B3")
    return _excel_response(wb, "plantilla_formas_farmaceuticas.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Importación desde Excel
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/productos/importar")
async def importar_productos(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")

    hacer_backup(etiqueta="pre_importacion")
    content = await file.read()
    try:
        headers, filas = _leer_excel(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    requeridos = {"codigo", "descripcion", "unidad"}
    faltantes_col = requeridos - set(headers)
    if faltantes_col:
        raise HTTPException(status_code=422,
            detail=f"Columnas faltantes en el archivo: {', '.join(faltantes_col)}")

    creados = actualizados = 0
    errores = []

    for i, row in enumerate(filas, start=2):
        if all(v is None for v in row):
            continue
        sp = db.begin_nested()
        try:
            codigo = _cel(row, headers, "codigo").upper()
            desc   = _cel(row, headers, "descripcion")
            unidad = _cel(row, headers, "unidad").upper()
            forma  = _cel(row, headers, "forma_farmaceutica")

            if not codigo: raise ValueError("El código está vacío.")
            if not desc:   raise ValueError("La descripción está vacía.")
            if unidad not in UNIDADES_VALIDAS:
                raise ValueError(f"Unidad '{unidad}' inválida. Usar UN, KG o L.")

            existente = db.query(ProductoTerminado).filter(
                ProductoTerminado.codigo == codigo).first()
            if existente:
                existente.descripcion = desc
                existente.unidad = UnidadMedida(unidad)
                existente.forma_farmaceutica = forma or None
                actualizados += 1
            else:
                db.add(ProductoTerminado(
                    codigo=codigo, descripcion=desc,
                    unidad=UnidadMedida(unidad),
                    forma_farmaceutica=forma or None,
                    activo=True,
                ))
                creados += 1
            sp.commit()
        except Exception as e:
            sp.rollback()
            errores.append({"fila": i, "error": str(e)})

    db.commit()
    return {"creados": creados, "actualizados": actualizados, "errores": errores}


def _resolver_forma(db: Session, nombre: str):
    """Busca la FormaFarmaceutica por nombre exacto (case-insensitive). Devuelve (id, nombre) o (None, nombre)."""
    if not nombre:
        return None, None
    ff = db.query(FormaFarmaceutica).filter(
        func.lower(FormaFarmaceutica.nombre) == nombre.lower()
    ).first()
    return (ff.id, ff.nombre) if ff else (None, nombre)


@router.post("/productos/importar-comprimidos")
async def importar_productos_comprimidos(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    hacer_backup(etiqueta="pre_importacion")
    content = await file.read()
    try:
        headers, filas = _leer_excel(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    requeridos = {"codigo", "descripcion", "unidad", "forma_farmaceutica"}
    faltantes_col = requeridos - set(headers)
    if faltantes_col:
        raise HTTPException(status_code=422,
            detail=f"Columnas faltantes: {', '.join(faltantes_col)}")

    creados = actualizados = 0
    errores = []

    for i, row in enumerate(filas, start=2):
        if all(v is None for v in row):
            continue
        sp = db.begin_nested()
        try:
            codigo = _cel(row, headers, "codigo").upper()
            desc   = _cel(row, headers, "descripcion")
            unidad = _cel(row, headers, "unidad").upper()
            forma_nombre = _cel(row, headers, "forma_farmaceutica")

            if not codigo: raise ValueError("El código está vacío.")
            if not desc:   raise ValueError("La descripción está vacía.")
            if unidad not in UNIDADES_VALIDAS:
                raise ValueError(f"Unidad '{unidad}' inválida. Usar UN, KG o L.")

            fid, fnombre = _resolver_forma(db, forma_nombre)
            peso   = _cel_num(row, headers, "peso_comprimido_mg")
            cxb    = _cel_num(row, headers, "comprimidos_x_blister")
            bxpt   = _cel_num(row, headers, "blisters_x_pt")

            existente = db.query(ProductoTerminado).filter(
                ProductoTerminado.codigo == codigo).first()
            if existente:
                existente.descripcion = desc
                existente.unidad = UnidadMedida(unidad)
                existente.forma_farmaceutica = fnombre or None
                existente.forma_farmaceutica_id = fid
                existente.peso_comprimido = int(peso) if peso is not None else None
                existente.cantidad_comprimidos_x_blister = int(cxb) if cxb is not None else None
                existente.cantidad_blisters_x_pt = int(bxpt) if bxpt is not None else None
                actualizados += 1
            else:
                db.add(ProductoTerminado(
                    codigo=codigo, descripcion=desc,
                    unidad=UnidadMedida(unidad),
                    forma_farmaceutica=fnombre or None,
                    forma_farmaceutica_id=fid,
                    activo=True,
                    peso_comprimido=int(peso) if peso is not None else None,
                    cantidad_comprimidos_x_blister=int(cxb) if cxb is not None else None,
                    cantidad_blisters_x_pt=int(bxpt) if bxpt is not None else None,
                ))
                creados += 1
            sp.commit()
        except Exception as e:
            sp.rollback()
            errores.append({"fila": i, "error": str(e)})

    db.commit()
    return {"creados": creados, "actualizados": actualizados, "errores": errores}


@router.post("/productos/importar-liquidos")
async def importar_productos_liquidos(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    hacer_backup(etiqueta="pre_importacion")
    content = await file.read()
    try:
        headers, filas = _leer_excel(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    requeridos = {"codigo", "descripcion", "unidad", "forma_farmaceutica"}
    faltantes_col = requeridos - set(headers)
    if faltantes_col:
        raise HTTPException(status_code=422,
            detail=f"Columnas faltantes: {', '.join(faltantes_col)}")

    creados = actualizados = 0
    errores = []

    for i, row in enumerate(filas, start=2):
        if all(v is None for v in row):
            continue
        sp = db.begin_nested()
        try:
            codigo = _cel(row, headers, "codigo").upper()
            desc   = _cel(row, headers, "descripcion")
            unidad = _cel(row, headers, "unidad").upper()
            forma_nombre = _cel(row, headers, "forma_farmaceutica")

            if not codigo: raise ValueError("El código está vacío.")
            if not desc:   raise ValueError("La descripción está vacía.")
            if unidad not in UNIDADES_VALIDAS:
                raise ValueError(f"Unidad '{unidad}' inválida. Usar UN, KG o L.")

            fid, fnombre = _resolver_forma(db, forma_nombre)
            cgu  = _cel_num(row, headers, "granel_x_unidad")
            cupt = _cel_num(row, headers, "unidades_x_pt")

            existente = db.query(ProductoTerminado).filter(
                ProductoTerminado.codigo == codigo).first()
            if existente:
                existente.descripcion = desc
                existente.unidad = UnidadMedida(unidad)
                existente.forma_farmaceutica = fnombre or None
                existente.forma_farmaceutica_id = fid
                existente.cantidad_granel_x_unidad = round(cgu, 1) if cgu is not None else None
                existente.cantidad_unidades_x_pt = int(cupt) if cupt is not None else None
                actualizados += 1
            else:
                db.add(ProductoTerminado(
                    codigo=codigo, descripcion=desc,
                    unidad=UnidadMedida(unidad),
                    forma_farmaceutica=fnombre or None,
                    forma_farmaceutica_id=fid,
                    activo=True,
                    cantidad_granel_x_unidad=round(cgu, 1) if cgu is not None else None,
                    cantidad_unidades_x_pt=int(cupt) if cupt is not None else None,
                ))
                creados += 1
            sp.commit()
        except Exception as e:
            sp.rollback()
            errores.append({"fila": i, "error": str(e)})

    db.commit()
    return {"creados": creados, "actualizados": actualizados, "errores": errores}


@router.post("/materias-primas/importar")
async def importar_mp(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")

    hacer_backup(etiqueta="pre_importacion")
    content = await file.read()
    try:
        headers, filas = _leer_excel(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    for col in ("codigo", "descripcion", "unidad"):
        if col not in headers:
            raise HTTPException(status_code=422, detail=f"Columna faltante: '{col}'")

    creados = actualizados = 0
    errores = []

    for i, row in enumerate(filas, start=2):
        if all(v is None for v in row):
            continue
        sp = db.begin_nested()
        try:
            codigo = _cel(row, headers, "codigo").upper()
            desc   = _cel(row, headers, "descripcion")
            unidad = _cel(row, headers, "unidad").upper()

            if not codigo: raise ValueError("El código está vacío.")
            if not desc:   raise ValueError("La descripción está vacía.")
            if unidad not in UNIDADES_VALIDAS_MP:
                raise ValueError(f"Unidad '{unidad}' inválida. Usar G, KG o L.")

            condicion = _cel(row, headers, "condicion").strip() or None
            if condicion and condicion not in CONDICIONES_VALIDAS:
                raise ValueError(f"Condición '{condicion}' inválida. Usar Activo o Excipiente.")

            existente = db.query(MateriaPrima).filter(MateriaPrima.codigo == codigo).first()
            if existente:
                existente.descripcion = desc
                existente.unidad = UnidadMedida(unidad)
                existente.condicion = condicion
                actualizados += 1
            else:
                db.add(MateriaPrima(codigo=codigo, descripcion=desc,
                                    unidad=UnidadMedida(unidad), condicion=condicion, activo=True))
                creados += 1
            sp.commit()
        except Exception as e:
            sp.rollback()
            errores.append({"fila": i, "error": str(e)})

    db.commit()
    return {"creados": creados, "actualizados": actualizados, "errores": errores}


@router.post("/materiales-empaque/importar")
async def importar_me(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")

    hacer_backup(etiqueta="pre_importacion")
    content = await file.read()
    try:
        headers, filas = _leer_excel(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    for col in ("codigo", "descripcion", "unidad"):
        if col not in headers:
            raise HTTPException(status_code=422, detail=f"Columna faltante: '{col}'")

    creados = actualizados = 0
    errores = []

    for i, row in enumerate(filas, start=2):
        if all(v is None for v in row):
            continue
        sp = db.begin_nested()
        try:
            codigo = _cel(row, headers, "codigo").upper()
            desc   = _cel(row, headers, "descripcion")
            unidad = _cel(row, headers, "unidad").upper()

            if not codigo: raise ValueError("El código está vacío.")
            if not desc:   raise ValueError("La descripción está vacía.")
            if unidad not in UNIDADES_VALIDAS_ME:
                raise ValueError(f"Unidad '{unidad}' inválida. Usar UN o KG.")
            clasificacion = _cel(row, headers, "clasificacion").strip() or None

            existente = db.query(MaterialEmpaque).filter(MaterialEmpaque.codigo == codigo).first()
            if existente:
                existente.descripcion = desc
                existente.unidad = UnidadMedida(unidad)
                existente.clasificacion = clasificacion
                actualizados += 1
            else:
                db.add(MaterialEmpaque(codigo=codigo, descripcion=desc,
                                       unidad=UnidadMedida(unidad),
                                       clasificacion=clasificacion, activo=True))
                creados += 1
            sp.commit()
        except Exception as e:
            sp.rollback()
            errores.append({"fila": i, "error": str(e)})

    db.commit()
    return {"creados": creados, "actualizados": actualizados, "errores": errores}


@router.post("/formas-farmaceuticas/importar")
async def importar_ff(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")

    hacer_backup(etiqueta="pre_importacion")
    content = await file.read()
    try:
        headers, filas = _leer_excel(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    if "nombre" not in headers:
        raise HTTPException(status_code=422, detail="Columna faltante: 'nombre'")

    creados = actualizados = 0
    errores = []

    for i, row in enumerate(filas, start=2):
        if all(v is None for v in row):
            continue
        sp = db.begin_nested()
        try:
            nombre     = _cel(row, headers, "nombre")
            etapas_raw = _cel(row, headers, "etapas")
            unidad_raw = _cel(row, headers, "unidad") if "unidad" in headers else None
            unidad_ff  = unidad_raw.upper() if unidad_raw else None

            if not nombre:
                raise ValueError("El nombre está vacío.")

            existente = db.query(FormaFarmaceutica).filter(
                FormaFarmaceutica.nombre == nombre).first()

            if existente:
                ff = existente
                if unidad_ff is not None:
                    ff.unidad = unidad_ff or None
                actualizados += 1
            else:
                ff = FormaFarmaceutica(nombre=nombre, activo=True, unidad=unidad_ff or None)
                db.add(ff)
                db.flush()  # necesario para obtener ff.id antes de agregar etapas
                creados += 1

            if etapas_raw:
                nombres_etapa = [e.strip() for e in etapas_raw.split(",") if e.strip()]
                existentes_nombres = {
                    e.nombre for e in db.query(EtapaProduccion).filter(
                        EtapaProduccion.forma_farmaceutica_id == ff.id
                    ).all()
                }
                orden_actual = db.query(EtapaProduccion).filter(
                    EtapaProduccion.forma_farmaceutica_id == ff.id
                ).count()
                for nombre_etapa in nombres_etapa:
                    if nombre_etapa not in existentes_nombres:
                        orden_actual += 1
                        db.add(EtapaProduccion(
                            forma_farmaceutica_id=ff.id,
                            orden=orden_actual,
                            nombre=nombre_etapa,
                            activo=True,
                        ))
            sp.commit()
        except Exception as e:
            sp.rollback()
            errores.append({"fila": i, "error": str(e)})

    db.commit()
    return {"creados": creados, "actualizados": actualizados, "errores": errores}
