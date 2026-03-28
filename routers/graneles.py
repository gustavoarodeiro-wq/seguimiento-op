from fastapi import APIRouter, Request, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db, Granel, ProductoTerminado, UnidadMedida
from routers.auth import require_auth

router = APIRouter(prefix="/api")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _granel_dict(g: Granel, total_productos: int = 0) -> dict:
    return {
        "id": g.id,
        "codigo": g.codigo,
        "descripcion": g.descripcion,
        "unidad": g.unidad.value,
        "activo": g.activo,
        "total_productos": total_productos,
    }


def _get_granel(db: Session, granel_id: int) -> Granel:
    g = db.query(Granel).filter(Granel.id == granel_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Granel no encontrado.")
    return g


def _excel_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _estilo_encabezado(ws, cols: list):
    fill = PatternFill("solid", fgColor="1D4ED8")
    font = Font(bold=True, color="FFFFFF")
    for i, nombre in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=nombre)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(nombre) + 4, 16)


def _leer_excel(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo está vacío.")
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def _cel(row, headers, col) -> str:
    try:
        idx = headers.index(col)
        v = row[idx]
        return str(v).strip() if v is not None else ""
    except (ValueError, IndexError):
        return ""


# ─────────────────────────────────────────────────────────────────────────────
# Excel plantilla e importación  (must be BEFORE parameterized routes)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/graneles/plantilla")
async def plantilla_graneles(user: dict = Depends(require_auth)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Graneles"
    cols = ["Código", "Descripción", "Unidad"]
    _estilo_encabezado(ws, cols)
    ws.append(["GR-001", "Doxiciclina 200mg Granel", "KG"])
    nota = ws.cell(row=3, column=1, value="* unidad: UN, KG, L o G")
    nota.font = Font(italic=True, color="6B7280")
    ws.merge_cells("A3:C3")
    return _excel_response(wb, "plantilla_graneles.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Graneles CRUD
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/graneles")
async def api_list_graneles(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    graneles = db.query(Granel).order_by(Granel.codigo).all()
    result = []
    for g in graneles:
        total = db.query(ProductoTerminado).filter(
            ProductoTerminado.granel_id == g.id
        ).count()
        result.append(_granel_dict(g, total))
    return result


@router.post("/graneles")
async def api_create_granel(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    body = await request.json()
    codigo = body.get("codigo", "").strip().upper()
    if not codigo:
        raise HTTPException(status_code=422, detail="El código es obligatorio.")
    if db.query(Granel).filter(Granel.codigo == codigo).first():
        raise HTTPException(status_code=409, detail=f"Ya existe el código '{codigo}'.")
    descripcion = body.get("descripcion", "").strip()
    if not descripcion:
        raise HTTPException(status_code=422, detail="La descripción es obligatoria.")
    try:
        unidad = UnidadMedida(body.get("unidad", "KG").strip().upper())
    except ValueError:
        raise HTTPException(status_code=422, detail="Unidad inválida. Usar UN, KG, L o G.")
    g = Granel(codigo=codigo, descripcion=descripcion, unidad=unidad, activo=True)
    db.add(g)
    db.commit()
    db.refresh(g)
    return _granel_dict(g, 0)


@router.post("/graneles/importar")
async def importar_graneles(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")

    content = await file.read()
    try:
        headers, filas = _leer_excel(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    # Normalize headers: strip and lowercase for matching
    headers_norm = [h.strip().lower() for h in headers]

    def _get_col(row, *names):
        for name in names:
            for i, h in enumerate(headers_norm):
                if h == name and i < len(row):
                    v = row[i]
                    return str(v).strip() if v is not None else ""
        return ""

    importados = 0
    errores = []

    for i, row in enumerate(filas, start=2):
        if all(v is None for v in row):
            continue
        sp = db.begin_nested()
        try:
            codigo = _get_col(row, "código", "codigo").upper()
            desc = _get_col(row, "descripción", "descripcion")
            unidad_str = _get_col(row, "unidad").upper()

            if not codigo:
                raise ValueError("El código está vacío.")
            if not desc:
                raise ValueError("La descripción está vacía.")
            try:
                unidad = UnidadMedida(unidad_str)
            except ValueError:
                raise ValueError(f"Unidad '{unidad_str}' inválida. Usar UN, KG, L o G.")

            existente = db.query(Granel).filter(Granel.codigo == codigo).first()
            if existente:
                existente.descripcion = desc
                existente.unidad = unidad
            else:
                db.add(Granel(codigo=codigo, descripcion=desc, unidad=unidad, activo=True))
            importados += 1
            sp.commit()
        except Exception as e:
            sp.rollback()
            errores.append({"fila": i, "error": str(e)})

    db.commit()
    return {"importados": importados, "errores": errores, "total": importados + len(errores)}


@router.put("/graneles/{granel_id}")
async def api_update_granel(
    granel_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    g = _get_granel(db, granel_id)
    body = await request.json()
    if "descripcion" in body:
        g.descripcion = body["descripcion"].strip()
    if "unidad" in body:
        try:
            g.unidad = UnidadMedida(body["unidad"].strip().upper())
        except ValueError:
            raise HTTPException(status_code=422, detail="Unidad inválida. Usar UN, KG, L o G.")
    db.commit()
    db.refresh(g)
    total = db.query(ProductoTerminado).filter(ProductoTerminado.granel_id == g.id).count()
    return _granel_dict(g, total)


@router.delete("/graneles/{granel_id}")
async def api_delete_granel(
    granel_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    g = _get_granel(db, granel_id)
    total = db.query(ProductoTerminado).filter(ProductoTerminado.granel_id == g.id).count()
    if total > 0:
        raise HTTPException(
            status_code=409,
            detail=f"No se puede eliminar: hay {total} producto(s) asociado(s) a este granel.",
        )
    db.delete(g)
    db.commit()
    return {"ok": True}


@router.patch("/graneles/{granel_id}/toggle")
async def api_toggle_granel(
    granel_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores.")
    g = _get_granel(db, granel_id)
    g.activo = not g.activo
    db.commit()
    return {"id": g.id, "activo": g.activo}


# ─────────────────────────────────────────────────────────────────────────────
# Asociación producto ↔ granel
# ─────────────────────────────────────────────────────────────────────────────

@router.patch("/productos/{producto_id}/granel")
async def api_set_granel_producto(
    producto_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    pt = db.query(ProductoTerminado).filter(ProductoTerminado.id == producto_id).first()
    if not pt:
        raise HTTPException(status_code=404, detail="Producto no encontrado.")
    body = await request.json()
    granel_id = body.get("granel_id")
    cantidad_granel = body.get("cantidad_granel")
    if granel_id is not None:
        g = db.query(Granel).filter(Granel.id == granel_id).first()
        if not g:
            raise HTTPException(status_code=404, detail="Granel no encontrado.")
    pt.granel_id = granel_id
    pt.cantidad_granel = cantidad_granel
    db.commit()
    db.refresh(pt)
    return {
        "id": pt.id,
        "granel_id": pt.granel_id,
        "cantidad_granel": pt.cantidad_granel,
        "granel_codigo": pt.granel.codigo if pt.granel else None,
    }


@router.delete("/productos/{producto_id}/granel")
async def api_remove_granel_producto(
    producto_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos.")
    pt = db.query(ProductoTerminado).filter(ProductoTerminado.id == producto_id).first()
    if not pt:
        raise HTTPException(status_code=404, detail="Producto no encontrado.")
    pt.granel_id = None
    pt.cantidad_granel = None
    db.commit()
    return {"id": pt.id, "granel_id": None, "cantidad_granel": None}
