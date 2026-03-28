from fastapi import APIRouter, Request, Depends, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
import io

from database import get_db, Formula, FormulaComponente, MateriaPrima, MaterialEmpaque, TipoFaltante
from routers.auth import require_auth

router = APIRouter()
templates = Jinja2Templates(directory="templates")
templates.env.cache = None


# ── Rutas fijas ANTES de las paramétricas ─────────────────────────────────────

@router.get("/api/formulas/plantilla")
async def api_plantilla_formulas(
    request: Request,
    user: dict = Depends(require_auth),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fórmulas"
    headers = ["Código PT", "Código Componente", "Cantidad por Unidad", "Unidad"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.append(["PROD-001", "MP-001", 100, "G"])
    ws.append(["PROD-001", "ME-001", 1, "UN"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_formulas.xlsx"},
    )


@router.post("/api/formulas/importar")
async def api_importar_formulas(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] not in ("admin", "operador"):
        raise HTTPException(403, "Sin permisos.")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(422, "El archivo debe ser .xlsx o .xls.")
    try:
        import openpyxl
        contenido = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(422, f"No se pudo leer el archivo: {e}")

    from database import ProductoTerminado
    mp_desc = {r.codigo: r.descripcion for r in db.query(MateriaPrima).all()}
    me_desc = {r.codigo: r.descripcion for r in db.query(MaterialEmpaque).all()}
    pt_desc = {r.codigo: r.descripcion for r in db.query(ProductoTerminado).all()}
    mp_codigos = set(mp_desc)
    me_codigos = set(me_desc)

    ok = 0
    errores = []
    formulas_cache = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        pt_raw      = str(row[0]).strip() if row[0] else ""
        comp_codigo = str(row[1]).strip().upper() if row[1] else ""
        cant_raw    = row[2]
        unidad_raw  = str(row[3]).strip().upper() if row[3] else ""

        if not pt_raw or not comp_codigo:
            errores.append(f"Fila {row_num}: producto o código vacío.")
            continue
        try:
            cantidad = float(cant_raw)
        except (TypeError, ValueError):
            errores.append(f"Fila {row_num}: cantidad inválida '{cant_raw}'.")
            continue
        if not unidad_raw:
            errores.append(f"Fila {row_num}: unidad vacía.")
            continue

        if comp_codigo in mp_codigos:
            tipo = TipoFaltante.MP
            comp_desc = mp_desc[comp_codigo]
        elif comp_codigo in me_codigos:
            tipo = TipoFaltante.ME
            comp_desc = me_desc[comp_codigo]
        else:
            errores.append(f"Fila {row_num}: '{comp_codigo}' no encontrado en MP ni ME.")
            continue

        pt_codigo = pt_raw.strip().upper()
        if pt_codigo not in formulas_cache:
            formula = db.query(Formula).filter(Formula.producto_codigo == pt_codigo).first()
            if not formula:
                # Buscar descripción en maestros según prefijo
                if pt_codigo.startswith("PT"):
                    pt_descripcion = pt_desc.get(pt_codigo, pt_raw)
                elif pt_codigo.startswith("MP"):
                    pt_descripcion = mp_desc.get(pt_codigo, pt_raw)
                elif pt_codigo.startswith("ME"):
                    pt_descripcion = me_desc.get(pt_codigo, pt_raw)
                else:
                    pt_descripcion = pt_raw
                formula = Formula(producto_codigo=pt_codigo, producto_descripcion=pt_descripcion)
                db.add(formula)
                db.flush()
            elif formula.producto_descripcion == formula.producto_codigo:
                # Corregir descripción si quedó igual al código
                if pt_codigo.startswith("PT"):
                    formula.producto_descripcion = pt_desc.get(pt_codigo, formula.producto_descripcion)
                elif pt_codigo.startswith("MP"):
                    formula.producto_descripcion = mp_desc.get(pt_codigo, formula.producto_descripcion)
                elif pt_codigo.startswith("ME"):
                    formula.producto_descripcion = me_desc.get(pt_codigo, formula.producto_descripcion)
            formulas_cache[pt_codigo] = formula
        else:
            formula = formulas_cache[pt_codigo]

        existe = db.query(FormulaComponente).filter(
            FormulaComponente.formula_id == formula.id,
            FormulaComponente.componente_codigo == comp_codigo,
        ).first()
        if existe:
            existe.cantidad = cantidad
            existe.unidad   = unidad_raw
            existe.componente_descripcion = comp_desc
        else:
            db.add(FormulaComponente(
                formula_id=formula.id, tipo=tipo,
                componente_codigo=comp_codigo, componente_descripcion=comp_desc,
                cantidad=cantidad, unidad=unidad_raw,
            ))
        ok += 1

    db.commit()
    return {"importados": ok, "errores": errores, "formulas_cargadas": len(formulas_cache)}


# ── Rutas paramétricas ────────────────────────────────────────────────────────

@router.get("/api/formulas")
async def api_list_formulas(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    formulas = db.query(Formula).order_by(Formula.producto_codigo).all()
    return [
        {**_formula_dict(f), "total_componentes": len(f.componentes)}
        for f in formulas
    ]


@router.get("/api/formulas-detalle/{formula_id}")
async def api_detalle_formula(
    formula_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    f = db.query(Formula).filter(Formula.id == formula_id).first()
    if not f:
        raise HTTPException(404, "Fórmula no encontrada.")
    return {**_formula_dict(f), "componentes": [_comp_dict(c) for c in f.componentes]}


@router.delete("/api/formulas/{formula_id}")
async def api_eliminar_formula(
    formula_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(403, "Solo el administrador puede eliminar fórmulas.")
    f = db.query(Formula).filter(Formula.id == formula_id).first()
    if not f:
        raise HTTPException(404, "Fórmula no encontrada.")
    db.delete(f)
    db.commit()
    return {"ok": True}


@router.get("/api/formulas/{producto_codigo}")
async def api_get_formula(
    producto_codigo: str,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    formula = db.query(Formula).filter(Formula.producto_codigo == producto_codigo).first()
    if not formula:
        return {"tiene_formula": False, "componentes": []}
    return {
        "tiene_formula": True,
        "formula_id": formula.id,
        "producto_codigo": formula.producto_codigo,
        "producto_descripcion": formula.producto_descripcion,
        "componentes": [_comp_dict(c) for c in formula.componentes],
    }


# ── Activar / desactivar fórmula ─────────────────────────────────────────────

@router.patch("/api/formulas/{formula_id}/toggle")
async def api_toggle_formula(
    formula_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(403, "Solo el administrador puede activar/desactivar fórmulas.")
    f = db.query(Formula).filter(Formula.id == formula_id).first()
    if not f:
        raise HTTPException(404, "Fórmula no encontrada.")
    f.activo = not f.activo
    db.commit()
    return {"id": f.id, "activo": f.activo}


# ── ABM de componentes ────────────────────────────────────────────────────────

@router.post("/api/formulas/{formula_id}/componentes")
async def api_agregar_componente(
    formula_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(403, "Solo el administrador puede modificar fórmulas.")
    f = db.query(Formula).filter(Formula.id == formula_id).first()
    if not f:
        raise HTTPException(404, "Fórmula no encontrada.")
    body = await request.json()
    comp_codigo = body.get("componente_codigo", "").strip().upper()
    if not comp_codigo:
        raise HTTPException(422, "El código del componente es obligatorio.")

    # Resolver tipo y descripción desde maestros
    mp = db.query(MateriaPrima).filter(MateriaPrima.codigo == comp_codigo).first()
    me = db.query(MaterialEmpaque).filter(MaterialEmpaque.codigo == comp_codigo).first()
    if mp:
        tipo = TipoFaltante.MP
        comp_desc = mp.descripcion
    elif me:
        tipo = TipoFaltante.ME
        comp_desc = me.descripcion
    else:
        raise HTTPException(422, f"Código '{comp_codigo}' no encontrado en MP ni ME.")

    existe = db.query(FormulaComponente).filter(
        FormulaComponente.formula_id == formula_id,
        FormulaComponente.componente_codigo == comp_codigo,
    ).first()
    if existe:
        raise HTTPException(409, f"El componente '{comp_codigo}' ya está en esta fórmula.")

    comp = FormulaComponente(
        formula_id=formula_id,
        tipo=tipo,
        componente_codigo=comp_codigo,
        componente_descripcion=comp_desc,
        cantidad=float(body.get("cantidad", 0)),
        unidad=body.get("unidad", "").strip().upper(),
    )
    db.add(comp)
    db.commit()
    db.refresh(comp)
    return _comp_dict(comp)


@router.patch("/api/formula-componentes/{comp_id}")
async def api_editar_componente(
    comp_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(403, "Solo el administrador puede modificar fórmulas.")
    comp = db.query(FormulaComponente).filter(FormulaComponente.id == comp_id).first()
    if not comp:
        raise HTTPException(404, "Componente no encontrado.")
    body = await request.json()
    if "cantidad" in body:
        comp.cantidad = float(body["cantidad"])
    if "unidad" in body:
        comp.unidad = body["unidad"].strip().upper()
    db.commit()
    return _comp_dict(comp)


@router.delete("/api/formula-componentes/{comp_id}")
async def api_eliminar_componente(
    comp_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    if user["rol"] != "admin":
        raise HTTPException(403, "Solo el administrador puede modificar fórmulas.")
    comp = db.query(FormulaComponente).filter(FormulaComponente.id == comp_id).first()
    if not comp:
        raise HTTPException(404, "Componente no encontrado.")
    db.delete(comp)
    db.commit()
    return {"ok": True}


# ── Vista HTML ────────────────────────────────────────────────────────────────

@router.get("/formulas", response_class=HTMLResponse)
async def page_formulas(
    request: Request,
    user: dict = Depends(require_auth),
):
    return templates.TemplateResponse(request, "maestros/formulas.html", {"user": user})


# ── Helpers ───────────────────────────────────────────────────────────────────

def _formula_dict(f: Formula) -> dict:
    return {
        "id": f.id,
        "producto_codigo": f.producto_codigo,
        "producto_descripcion": f.producto_descripcion,
        "activo": f.activo,
    }


def _comp_dict(c: FormulaComponente) -> dict:
    return {
        "id": c.id,
        "tipo": c.tipo.value,
        "componente_codigo": c.componente_codigo,
        "componente_descripcion": c.componente_descripcion,
        "cantidad": c.cantidad,
        "unidad": c.unidad,
    }
