"""
Exporta todos los maestros a archivos Excel para importar en producción.
Uso: python exportar_maestros.py
"""
import os
os.environ.pop("DATABASE_URL", None)  # forzar SQLite local

from database import SessionLocal
from database import (
    FormaFarmaceutica, EtapaProduccion,
    MateriaPrima, MaterialEmpaque, ProductoTerminado, Granel,
    Formula, FormulaComponente
)
import openpyxl

db = SessionLocal()

# ── Formas Farmacéuticas ───────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Formas"
ws.append(["nombre", "unidad", "etapas", "activo"])

formas = db.query(FormaFarmaceutica).order_by(FormaFarmaceutica.nombre).all()
for f in formas:
    etapas_activas = [e.nombre for e in sorted(f.etapas, key=lambda x: x.orden) if e.activo]
    ws.append([f.nombre, f.unidad or "", ", ".join(etapas_activas), "SI" if f.activo else "NO"])

wb.save("export_formas_farmaceuticas.xlsx")
print(f"✓ export_formas_farmaceuticas.xlsx — {len(formas)} registros")

# ── Materias Primas ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MateriasPrimas"
ws.append(["codigo", "descripcion", "unidad", "condicion", "activo"])

items = db.query(MateriaPrima).order_by(MateriaPrima.codigo).all()
for i in items:
    ws.append([i.codigo, i.descripcion, i.unidad.value, i.condicion or "", "SI" if i.activo else "NO"])

wb.save("export_materias_primas.xlsx")
print(f"✓ export_materias_primas.xlsx — {len(items)} registros")

# ── Materiales de Empaque ──────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MaterialesEmpaque"
ws.append(["codigo", "descripcion", "unidad", "clasificacion", "activo"])

items = db.query(MaterialEmpaque).order_by(MaterialEmpaque.codigo).all()
for i in items:
    ws.append([i.codigo, i.descripcion, i.unidad.value, i.clasificacion or "", "SI" if i.activo else "NO"])

wb.save("export_materiales_empaque.xlsx")
print(f"✓ export_materiales_empaque.xlsx — {len(items)} registros")

# ── Productos Terminados ───────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos"
ws.append(["codigo", "descripcion", "unidad", "forma_farmaceutica", "activo"])

items = db.query(ProductoTerminado).order_by(ProductoTerminado.codigo).all()
for i in items:
    ws.append([i.codigo, i.descripcion, i.unidad.value, i.forma_farmaceutica or "", "SI" if i.activo else "NO"])

wb.save("export_productos.xlsx")
print(f"✓ export_productos.xlsx — {len(items)} registros")

# ── Graneles ───────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Graneles"
ws.append(["codigo", "descripcion", "unidad", "activo"])

items = db.query(Granel).order_by(Granel.codigo).all()
for i in items:
    ws.append([i.codigo, i.descripcion, i.unidad.value, "SI" if i.activo else "NO"])

wb.save("export_graneles.xlsx")
print(f"✓ export_graneles.xlsx — {len(items)} registros")

# ── Fórmulas ───────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Formulas"
ws.append(["producto_codigo", "producto_descripcion", "tipo", "componente_codigo", "componente_descripcion", "cantidad", "unidad"])

formulas = db.query(Formula).order_by(Formula.producto_codigo).all()
filas = 0
for f in formulas:
    for c in f.componentes:
        ws.append([f.producto_codigo, f.producto_descripcion, c.tipo.value, c.componente_codigo, c.componente_descripcion, c.cantidad, c.unidad])
        filas += 1

wb.save("export_formulas.xlsx")
print(f"✓ export_formulas.xlsx — {len(formulas)} fórmulas, {filas} componentes")

# ── Productos Comprimidos ──────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comprimidos"
ws.append(["codigo", "descripcion", "unidad", "forma_farmaceutica",
           "peso_comprimido_mg", "comprimidos_x_blister", "blisters_x_pt"])

items = (
    db.query(ProductoTerminado)
    .join(FormaFarmaceutica, ProductoTerminado.forma_farmaceutica_id == FormaFarmaceutica.id, isouter=True)
    .filter(FormaFarmaceutica.nombre.ilike("%comprimido%"))
    .order_by(ProductoTerminado.codigo)
    .all()
)
for p in items:
    ws.append([p.codigo, p.descripcion, p.unidad.value, p.forma_farmaceutica or "",
               p.peso_comprimido, p.cantidad_comprimidos_x_blister, p.cantidad_blisters_x_pt])

wb.save("export_productos_comprimidos.xlsx")
print(f"✓ export_productos_comprimidos.xlsx — {len(items)} registros")

# ── Productos Líquidos/Sólidos ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Liquidos y Solidos"
ws.append(["codigo", "descripcion", "unidad", "forma_farmaceutica",
           "granel_x_unidad", "unidades_x_pt"])

items = (
    db.query(ProductoTerminado)
    .join(FormaFarmaceutica, ProductoTerminado.forma_farmaceutica_id == FormaFarmaceutica.id, isouter=True)
    .filter(
        FormaFarmaceutica.unidad.isnot(None),
        ~FormaFarmaceutica.nombre.ilike("%comprimido%"),
    )
    .order_by(ProductoTerminado.codigo)
    .all()
)
for p in items:
    ws.append([p.codigo, p.descripcion, p.unidad.value, p.forma_farmaceutica or "",
               p.cantidad_granel_x_unidad, p.cantidad_unidades_x_pt])

wb.save("export_productos_liquidos.xlsx")
print(f"✓ export_productos_liquidos.xlsx — {len(items)} registros")

db.close()
print("\nListo. Los archivos Excel están en la carpeta del proyecto.")
